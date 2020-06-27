#!/usr/bin/env python3
# coding:utf-8
from copy import copy
from tempfile import NamedTemporaryFile

from openpyxl import load_workbook
from openpyxl.styles import Border, Alignment, Font

from nemo.common.utils.assertinfoparser import AssertInfoParser
from nemo.core.database.ip import Ip
from nemo.core.database.domain import Domain
from nemo.core.database.port import Port
from nemo.core.database.organization import Organization
from nemo.core.database.attr import DomainAttr, PortAttr

template_path = 'nemo/web/templates'
template_file_domain = '{}/domain-export.xlsx'.format(template_path)
template_file_ip = '{}/ip-export.xlsx'.format(template_path)


def _get_domains(org_id, domain_address, ip_address):
    '''获取域名
    '''
    domain_app = Domain()
    domain_attr_app = DomainAttr()
    org_app = Organization()
    api = AssertInfoParser()

    domain_list = []
    domains = domain_app.gets_by_org_domain_ip(
        org_id, domain_address, ip_address, page=1, rows_per_page=100000)
    if domains:
        for index, domain_row in enumerate(domains):
            ips = domain_attr_app.gets(
                query={'tag': 'A', 'r_id': domain_row['id']})
            domain_info = api.get_domain_info(domain_row['id'])
            domain_list.append({
                'id': domain_row['id'],
                "index": index+1,
                "domain": domain_row['domain'],
                "ip": ', '.join(set([ip_row['content'] for ip_row in ips])),
                "org_name": org_app.get(int(domain_row['org_id']))['org_name'] if domain_row['org_id'] else '',
                "create_time": str(domain_row['create_datetime']),
                "update_time": str(domain_row['update_datetime']),
                'port': ', '.join([str(x) for x in domain_info['port']]),
                'title': '\n'.join(domain_info['title']),
                'banner': '\n'.join(domain_info['banner'])
            })

    return domain_list


def _get_ips(org_id, ip_address, port):
    '''获取IP
    '''
    ip_table = Ip()
    aip = AssertInfoParser()

    ip_list = []
    ips = ip_table.gets_by_org_ip_port(
        org_id, ip_address, port, page=1, rows_per_page=100000)
    if ips:
        for i, ip_row in enumerate(ips):
            ip_info = aip.get_ip_info(ip_row['id'])
            ip_info.update(index= i+1)
            ip_list.append(ip_info)

    return ip_list


def _copy_cell_style(ws, src_row, dst_row, col_from, col_to):
    '''复制单元格格式
    '''
    for i in range(col_from, col_to+1):
        ws.cell(column=i, row=dst_row).border = copy(
            ws.cell(column=i, row=src_row).border)
        ws.cell(column=i, row=dst_row).font = copy(
            ws.cell(column=i, row=src_row).font)
        ws.cell(column=i, row=dst_row).fill = copy(
            ws.cell(column=i, row=src_row).fill)
        ws.cell(column=i, row=dst_row).alignment = copy(
            ws.cell(column=i, row=src_row).alignment)


def export_domains(org_id=None, domain_address=None, ip_address=None):
    '''导出域名为excel文件
    '''
    wb = load_workbook(template_file_domain)
    ws = wb.active
    domains = _get_domains(org_id, domain_address, ip_address)
    row_start = 2
    for domain in domains:
        _copy_cell_style(ws, 2, row_start, 1, 7)
        ws.cell(column=1, row=row_start, value="{0}".format(domain['index']))
        ws.cell(column=2, row=row_start, value="{0}".format(domain['domain']))
        ws.cell(column=3, row=row_start, value="{0}".format(domain['ip']))
        ws.cell(column=4, row=row_start, value="{0}".format(domain['port']))
        ws.cell(column=5, row=row_start, value="{0}".format(domain['title']))
        ws.cell(column=6, row=row_start, value="{0}".format(domain['banner']))
        row_start += 1

    with NamedTemporaryFile() as tmp:
        wb.save(tmp.name)
        tmp.seek(0)
        data = tmp.read()
        return data


def export_ips(org_id=None, ip_address=None, port=None):
    '''导出IP为excel文件
    '''
    wb = load_workbook(template_file_ip)
    ws = wb.active
    ips = _get_ips(org_id, ip_address, port)
    row_start = 2
    for ip in ips:
        merged_row_start = row_start
        _copy_cell_style(ws, 2, row_start, 1, 9)
        if ip['port_attr']:
            for port in ip['port_attr']:
                _copy_cell_style(ws, 2, row_start, 1, 9)
                ws.cell(column=5, row=row_start,
                        value="{0}".format(port['port']))
                ws.cell(column=6, row=row_start,
                        value="{0}".format(port['source']))
                ws.cell(column=7, row=row_start,
                        value="{0}".format(port['tag']))
                ws.cell(column=8, row=row_start,
                        value="{0}".format(port['content']))
                ws.cell(column=9, row=row_start,
                        value="{0}".format(port['update_datetime']))
                row_start += 1
        else:
            row_start += 1
        ws.merge_cells(start_row=merged_row_start, start_column=1, end_row=row_start-1, end_column=1)
        ws.merge_cells(start_row=merged_row_start, start_column=2, end_row=row_start-1, end_column=2)
        ws.merge_cells(start_row=merged_row_start, start_column=3, end_row=row_start-1, end_column=3)
        ws.merge_cells(start_row=merged_row_start, start_column=4, end_row=row_start-1, end_column=4)

        ws.cell(column=1, row=merged_row_start, value="{0}".format(ip['index']))
        ws.cell(column=2, row=merged_row_start, value="{0}".format(ip['ip']))
        ws.cell(column=3, row=merged_row_start,
                value="{0}".format('\n'.join(ip['domain'])))
        ws.cell(column=4, row=merged_row_start, value="{0}".format(
            ip['location'] if ip['location'] else ''))

    with NamedTemporaryFile() as tmp:
        wb.save(tmp.name)
        tmp.seek(0)
        data = tmp.read()
        return data
